import openpyxl
import webbrowser

def open_urls_in_chrome(excel_file_path):
    """Opens URLs from each sheet of an Excel file in separate Chrome windows on Linux."""

    workbook = openpyxl.load_workbook(excel_file_path)
    windows_opened = 0

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Determine column containing URLs based on header or pattern
        url_column = get_url_column(sheet)

        # Extract URLs from the column
        urls = get_urls_from_column(sheet, url_column)

        # Open URLs in a new Chrome window
        open_urls_in_window(urls)

        windows_opened += 1

    print(f"Opened {windows_opened} Chrome windows with URLs from {excel_file_path}")

def get_url_column(sheet):
    """ identifies the column containing URLs."""

    column_index = None

    # Check for explicit "URL" or "Link" header
    for col in sheet.iter_cols():
        if "URL" in col[0].value or "Link" in col[0].value:
            column_index = col[0].col_idx - 1  # Convert 1-based to 0-based indexing
            break

    # If not found, check for column with valid URLs based on a pattern
    if column_index is None:
        for col in sheet.iter_cols():
            for cell in col:
                if isinstance(cell.value, str) and is_valid_url(cell.value):
                    column_index = col[0].col_idx - 1
                    break  # Assume only one column contains valid URLs
            if column_index is not None:
                break

    return column_index

def get_urls_from_column(sheet, url_column):
    """Extracts URLs from the specified column."""

    urls = []
    for row in sheet.iter_rows(min_col=url_column + 1, min_row=2):  # Skip header row
        url = row[url_column].value
        print(url)
        if is_valid_url(url):
            urls.append(url)
    return urls

def is_valid_url(url):
    """Basic check for a valid URL format."""

    import re
    url_pattern = r"^(http|https)://\w+(\.\w+)*(/[^\s]*)*$"
    return bool(re.match(url_pattern, url))

def open_urls_in_window(urls):
    """Opens the given URLs in a new Chrome window."""

    try:
        # import os
        # chrome_path = "/usr/bin/google-chrome"  # Customize for your Chrome path
        # if not os.path.exists(chrome_path):
        #     raise ValueError("Chrome path not found")
        count = 0
        for url in urls:            
            if count == 0:
                webbrowser.get().open_new(url)
                count+=1
            else:
                webbrowser.get().open(url)   
    except Exception as e:
        print(f"Error opening Chrome window: {e}")

if __name__ == "__main__":
    excel_file_path = "links.xlsx"  # Replace with your file path
    open_urls_in_chrome(excel_file_path)